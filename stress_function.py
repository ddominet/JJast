import time
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook

def stress_function(num):
    return sum([i*j*k for i in range(num) for j in range(i) for k in range(j)])

inp = int(input("Enter a number: "))
czas = []

try:
    wb = Workbook()
except:
    wb = load_workbook("excel_data.xlsx")
sheet = wb.active

for i in range(inp):
    start = time.time()
    test = stress_function(i)
    czas.append(time.time() - start)

for i in range(inp):
    sheet.cell(row=i+1, column=1).value = i+1
    sheet.cell(row=i+1, column=2).value = czas[i]

wb.save("excel_data.xlsx")

dev_x = [i for i in range(inp)]
dev_y = [czas[i] for i in range(inp)]

plt.xlabel("Number of iterations")
plt.ylabel("Time in seconds")
plt.title("Time of compilation by number of iterations")

plt.plot(dev_x, dev_y)

plt.legend(["Stress Function"])

plt.show(block=True)
